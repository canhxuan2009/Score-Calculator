"""Read-only ingestion of the supported single-sheet workbook format."""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections.abc import Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Final, cast

from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from point_audit.domain import (
    DomainWarning,
    RawCell,
    RawWorkbookRow,
    ScoringPeriod,
    SourceColumn,
    WarningCode,
)
from point_audit.ingestion.errors import (
    HeaderAmbiguousError,
    HeaderNotFoundError,
    SourceWorkbookChangedError,
    WorkbookStructureError,
)
from point_audit.ingestion.models import (
    DetectedColumn,
    IngestedStudentRow,
    WorkbookIngestionResult,
)

_REQUIRED_COLUMNS: Final[frozenset[SourceColumn]] = frozenset(
    {
        SourceColumn.FULL_NAME,
        SourceColumn.POSITIVE_TOTAL,
        SourceColumn.NEGATIVE_TOTAL,
        SourceColumn.FINAL_TOTAL,
        SourceColumn.EVIDENCE,
    }
)

_HEADER_ALIASES: Final[dict[str, SourceColumn]] = {
    "tt": SourceColumn.SEQUENCE,
    "stt": SourceColumn.SEQUENCE,
    "hovaten": SourceColumn.FULL_NAME,
    "hoten": SourceColumn.FULL_NAME,
    "ngaysinh": SourceColumn.BIRTH_DATE,
    "sinhngay": SourceColumn.BIRTH_DATE,
    "to": SourceColumn.GROUP,
    "diemgoc": SourceColumn.BASE_SCORE,
    "diemcong": SourceColumn.POSITIVE_TOTAL,
    "diemtru": SourceColumn.NEGATIVE_TOTAL,
    "tong": SourceColumn.FINAL_TOTAL,
    "tongdiem": SourceColumn.FINAL_TOTAL,
    "hanhkiem": SourceColumn.CONDUCT,
    "minhchung": SourceColumn.EVIDENCE,
}

_STOP_MARKERS: Final[tuple[str, ...]] = ("tonghop", "thanhtichlop", "gvcn")
_PERIOD_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?P<name>đợt\s*[^:–—\-\n]+)?\s*:?[ \t]*"
    r"từ\s*(?P<start>\d{1,2}[./-]\d{1,2}[./-]\d{4})\s*"
    r"(?:–|—|-|đến)\s*(?P<end>\d{1,2}[./-]\d{1,2}[./-]\d{4})",
    flags=re.IGNORECASE,
)
CellLike = Cell | MergedCell


class WorkbookReader:
    """Inspect one workbook without mutating or saving it."""

    def read(self, source_path: str | Path) -> WorkbookIngestionResult:
        """Read the workbook, rejecting unsupported or ambiguous structure."""
        path = Path(source_path)
        source_hash = _sha256(path)
        keep_vba = path.suffix.casefold() == ".xlsm"
        formula_workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_vba=keep_vba,
        )
        cached_workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_vba=keep_vba,
        )
        try:
            if len(formula_workbook.worksheets) != 1:
                raise WorkbookStructureError("Workbook phải có đúng một sheet cần xử lý.")
            if len(cached_workbook.worksheets) != 1:
                raise WorkbookStructureError(
                    "Không thể đọc nhất quán sheet và giá trị công thức đã lưu."
                )

            formula_sheet = formula_workbook.worksheets[0]
            cached_sheet = cached_workbook.worksheets[0]
            header_row, columns = self._find_header(formula_sheet)
            scoring_period = self._find_scoring_period(formula_sheet, header_row)
            students, stopped_at, formulas_found, cached_found = self._read_students(
                formula_sheet=formula_sheet,
                cached_sheet=cached_sheet,
                source_hash=source_hash,
                header_row=header_row,
                columns=columns,
                workbook_epoch=formula_workbook.epoch,
            )
            sheet_name = formula_sheet.title
        finally:
            formula_workbook.close()
            cached_workbook.close()

        if _sha256(path) != source_hash:
            raise SourceWorkbookChangedError(
                "Workbook nguồn đã thay đổi trong lúc đọc; kết quả đã bị hủy."
            )

        return WorkbookIngestionResult(
            source_file_name=path.name,
            source_file_sha256=source_hash,
            sheet_name=sheet_name,
            header_row=header_row,
            columns=columns,
            scoring_period=scoring_period,
            students=students,
            stopped_at_row=stopped_at,
            formulas_found=formulas_found,
            cached_formula_values_found=cached_found,
        )

    def _find_header(
        self, worksheet: Worksheet
    ) -> tuple[int, tuple[DetectedColumn, ...]]:
        candidates: list[tuple[int, tuple[DetectedColumn, ...]]] = []
        for excel_row, row in enumerate(worksheet.iter_rows(), start=1):
            detected: dict[SourceColumn, DetectedColumn] = {}
            duplicate = False
            for cell in row:
                header_text = _raw_text(cell.value)
                if header_text is None:
                    continue
                source_column = _HEADER_ALIASES.get(_normalize(header_text))
                if source_column is None:
                    continue
                if not isinstance(cell.column, int):
                    continue
                if source_column in detected:
                    duplicate = True
                    break
                detected[source_column] = DetectedColumn(
                    source_column=source_column,
                    excel_column=cell.column,
                    source_column_name=header_text,
                )
            if not duplicate and _REQUIRED_COLUMNS.issubset(detected):
                ordered = tuple(sorted(detected.values(), key=lambda item: item.excel_column))
                candidates.append((excel_row, ordered))

        if not candidates:
            required = ", ".join(column.value for column in sorted(_REQUIRED_COLUMNS))
            raise HeaderNotFoundError(f"Không tìm thấy hàng tiêu đề chứa đủ: {required}.")
        if len(candidates) > 1:
            rows = ", ".join(str(candidate[0]) for candidate in candidates)
            raise HeaderAmbiguousError(f"Nhiều hàng có thể là tiêu đề: {rows}.")
        return candidates[0]

    def _find_scoring_period(
        self, worksheet: Worksheet, header_row: int
    ) -> ScoringPeriod | None:
        if header_row == 1:
            return None
        for row in worksheet.iter_rows(min_row=1, max_row=header_row - 1):
            for cell in row:
                text = _raw_text(cell.value)
                if text is None:
                    continue
                match = _PERIOD_PATTERN.search(text)
                if match is None:
                    continue
                starts_on = _parse_full_date(match.group("start"))
                ends_on = _parse_full_date(match.group("end"))
                name = (match.group("name") or "Khoảng chấm điểm").strip(" :")
                identity = f"{name}\x1f{starts_on.isoformat()}\x1f{ends_on.isoformat()}"
                period_id = f"period_{hashlib.sha256(identity.encode()).hexdigest()[:16]}"
                academic_year = (
                    str(starts_on.year)
                    if starts_on.year == ends_on.year
                    else f"{starts_on.year}-{ends_on.year}"
                )
                return ScoringPeriod(
                    period_id=period_id,
                    name=name,
                    starts_on=starts_on,
                    ends_on=ends_on,
                    academic_year_label=academic_year,
                )
        return None

    def _read_students(
        self,
        *,
        formula_sheet: Worksheet,
        cached_sheet: Worksheet,
        source_hash: str,
        header_row: int,
        columns: tuple[DetectedColumn, ...],
        workbook_epoch: datetime,
    ) -> tuple[tuple[IngestedStudentRow, ...], int | None, bool, bool]:
        by_source = {column.source_column: column for column in columns}
        students: list[IngestedStudentRow] = []
        stopped_at: int | None = None
        formulas_found = False
        cached_found = False

        formula_rows = formula_sheet.iter_rows(min_row=header_row + 1)
        cached_rows = cached_sheet.iter_rows(min_row=header_row + 1)
        for excel_row, (formula_cells, cached_cells) in enumerate(
            zip(formula_rows, cached_rows, strict=True),
            start=header_row + 1,
        ):
            if self._is_stop_row(formula_cells, by_source):
                stopped_at = excel_row
                break

            row_values = self._mapped_values(formula_cells, cached_cells, by_source)
            sequence_number = _parse_sequence(row_values.get(SourceColumn.SEQUENCE))
            full_name = _optional_text(row_values.get(SourceColumn.FULL_NAME))
            if sequence_number is None and full_name is None:
                continue

            raw_cells: list[RawCell] = []
            for column in columns:
                formula_cell = _cell_at(formula_cells, column.excel_column)
                cached_cell = _cell_at(cached_cells, column.excel_column)
                raw_cell = _build_raw_cell(
                    formula_cell=formula_cell,
                    cached_cell=cached_cell,
                    source_hash=source_hash,
                    sheet_name=formula_sheet.title,
                    excel_row=excel_row,
                    detected_column=column,
                )
                raw_cells.append(raw_cell)
                if raw_cell.formula is not None:
                    formulas_found = True
                    if raw_cell.cached_value_text is not None:
                        cached_found = True

            warnings: list[DomainWarning] = []
            if full_name is None:
                warnings.append(
                    _warning(
                        WarningCode.MISSING_REQUIRED_COLUMN,
                        "Dòng có TT dạng số nhưng thiếu Họ và tên.",
                    )
                )

            base_score = _parse_decimal_field(
                row_values.get(SourceColumn.BASE_SCORE), "Điểm gốc", warnings
            )
            positive = _parse_nonnegative_field(
                row_values.get(SourceColumn.POSITIVE_TOTAL), "Điểm cộng", warnings
            )
            negative = _parse_nonnegative_field(
                row_values.get(SourceColumn.NEGATIVE_TOTAL), "Điểm trừ", warnings
            )
            final = _parse_decimal_field(
                row_values.get(SourceColumn.FINAL_TOTAL), "Tổng", warnings
            )
            birth_value = row_values.get(SourceColumn.BIRTH_DATE)
            birth_date = _parse_birth_date(birth_value, workbook_epoch, warnings)

            if all(value is not None for value in (base_score, positive, negative, final)):
                assert base_score is not None
                assert positive is not None
                assert negative is not None
                assert final is not None
                if final != base_score + positive - negative:
                    warnings.append(
                        _warning(
                            WarningCode.ROW_FORMULA_MISMATCH,
                            "Tổng đã khai báo không bằng Điểm gốc + Điểm cộng - Điểm trừ.",
                        )
                    )

            raw_by_source = {cell.source_column: cell.raw_text for cell in raw_cells}
            students.append(
                IngestedStudentRow(
                    source_file_sha256=source_hash,
                    sheet_name=formula_sheet.title,
                    excel_row=excel_row,
                    sequence_raw=_empty_to_none(raw_by_source.get(SourceColumn.SEQUENCE)),
                    sequence_number=sequence_number,
                    full_name_raw=_empty_to_none(raw_by_source.get(SourceColumn.FULL_NAME)),
                    birth_date_raw=_empty_to_none(raw_by_source.get(SourceColumn.BIRTH_DATE)),
                    birth_date=birth_date,
                    group_raw=_empty_to_none(raw_by_source.get(SourceColumn.GROUP)),
                    base_score_raw=_empty_to_none(raw_by_source.get(SourceColumn.BASE_SCORE)),
                    base_score=base_score,
                    declared_positive_raw=_empty_to_none(
                        raw_by_source.get(SourceColumn.POSITIVE_TOTAL)
                    ),
                    declared_positive_total=positive,
                    declared_negative_raw=_empty_to_none(
                        raw_by_source.get(SourceColumn.NEGATIVE_TOTAL)
                    ),
                    declared_negative_total=negative,
                    declared_final_raw=_empty_to_none(raw_by_source.get(SourceColumn.FINAL_TOTAL)),
                    declared_final_total=final,
                    conduct_raw=_empty_to_none(raw_by_source.get(SourceColumn.CONDUCT)),
                    evidence_raw=raw_by_source.get(SourceColumn.EVIDENCE, ""),
                    raw_row=RawWorkbookRow(
                        source_file_sha256=source_hash,
                        sheet_name=formula_sheet.title,
                        excel_row=excel_row,
                        cells=tuple(raw_cells),
                    ),
                    warnings=tuple(warnings),
                )
            )

        return tuple(students), stopped_at, formulas_found, cached_found

    @staticmethod
    def _mapped_values(
        formula_cells: Sequence[CellLike],
        cached_cells: Sequence[CellLike],
        columns: dict[SourceColumn, DetectedColumn],
    ) -> dict[SourceColumn, Any]:
        values: dict[SourceColumn, Any] = {}
        for source_column, detected in columns.items():
            formula_cell = _cell_at(formula_cells, detected.excel_column)
            cached_cell = _cell_at(cached_cells, detected.excel_column)
            values[source_column] = _semantic_value(formula_cell, cached_cell)
        return values

    @staticmethod
    def _is_stop_row(
        row: Sequence[CellLike], columns: dict[SourceColumn, DetectedColumn]
    ) -> bool:
        inspected_columns = {1, 2, 3}
        for source_column in (SourceColumn.SEQUENCE, SourceColumn.FULL_NAME):
            detected = columns.get(source_column)
            if detected is not None:
                inspected_columns.add(detected.excel_column)
        for excel_column in inspected_columns:
            text = _raw_text(_cell_at(row, excel_column).value)
            if text is None:
                continue
            normalized = _normalize(text)
            if any(normalized.startswith(marker) for marker in _STOP_MARKERS):
                return True
        return False


def _build_raw_cell(
    *,
    formula_cell: CellLike,
    cached_cell: CellLike,
    source_hash: str,
    sheet_name: str,
    excel_row: int,
    detected_column: DetectedColumn,
) -> RawCell:
    formula = _formula_text(formula_cell)
    raw = formula if formula is not None else (_raw_text(formula_cell.value) or "")
    cached = _raw_text(cached_cell.value) if formula is not None else None
    return RawCell(
        source_file_sha256=source_hash,
        sheet_name=sheet_name,
        excel_row=excel_row,
        excel_column=detected_column.excel_column,
        source_column=detected_column.source_column,
        source_column_name=detected_column.source_column_name,
        raw_text=raw,
        formula=formula,
        cached_value_text=cached,
    )


def _semantic_value(formula_cell: CellLike, cached_cell: CellLike) -> Any:
    return cached_cell.value if _formula_text(formula_cell) is not None else formula_cell.value


def _formula_text(cell: CellLike) -> str | None:
    if cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    ):
        return str(cell.value)
    return None


def _cell_at(row: Sequence[CellLike], excel_column: int) -> CellLike:
    return row[excel_column - 1]


def _normalize(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.casefold().replace("đ", "d"))
    unaccented = "".join(
        character for character in decomposed if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", "", unaccented)


def _raw_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _optional_text(value: Any) -> str | None:
    return _empty_to_none(_raw_text(value))


def _empty_to_none(value: str | None) -> str | None:
    if value is None or not value.strip():
        return None
    return value


def _parse_sequence(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        decimal_value = Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return None
    if not decimal_value.is_finite() or decimal_value != decimal_value.to_integral_value():
        return None
    sequence = int(decimal_value)
    return sequence if sequence >= 0 else None


def _parse_decimal_field(
    value: Any, label: str, warnings: list[DomainWarning]
) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    decimal_value: Decimal | None
    if isinstance(value, bool):
        decimal_value = None
    else:
        try:
            decimal_value = Decimal(
                str(value).strip().replace(" ", "").replace(",", ".")
            )
        except (InvalidOperation, ValueError):
            decimal_value = None
    if decimal_value is None or not decimal_value.is_finite():
        warnings.append(
            _warning(WarningCode.INVALID_NUMBER, f"{label} không phải là số hữu hạn hợp lệ.")
        )
        return None
    return decimal_value


def _parse_nonnegative_field(
    value: Any, label: str, warnings: list[DomainWarning]
) -> Decimal | None:
    parsed = _parse_decimal_field(value, label, warnings)
    if parsed is not None and parsed < 0:
        warnings.append(_warning(WarningCode.INVALID_NUMBER, f"{label} không được là số âm."))
        return None
    return parsed


def _parse_birth_date(
    value: Any, workbook_epoch: datetime, warnings: list[DomainWarning]
) -> date | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, bool):
            raise ValueError
        if isinstance(value, (int, float, Decimal)):
            converted = from_excel(value, workbook_epoch)
            return converted.date() if isinstance(converted, datetime) else cast(date, converted)
        text = str(value).strip()
        if re.fullmatch(r"\d+(?:[.,]\d+)?", text):
            serial = Decimal(text.replace(",", "."))
            converted = from_excel(serial, workbook_epoch)
            return converted.date() if isinstance(converted, datetime) else cast(date, converted)
        return date_parser.parse(text, dayfirst=True, fuzzy=False).date()
    except (OverflowError, TypeError, ValueError):
        warnings.append(
            _warning(WarningCode.INVALID_DATE, "Ngày sinh không thể chuyển thành ngày hợp lệ.")
        )
        return None


def _parse_full_date(value: str) -> date:
    return date_parser.parse(value, dayfirst=True, fuzzy=False).date()


def _warning(code: WarningCode, message_vi: str) -> DomainWarning:
    return DomainWarning(code=code, message_vi=message_vi, blocking=False)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
