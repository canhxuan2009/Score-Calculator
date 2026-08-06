from __future__ import annotations

import hashlib
import re
import zipfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import from_excel

from point_audit.domain import SourceColumn, WarningCode
from point_audit.ingestion import (
    HeaderAmbiguousError,
    HeaderNotFoundError,
    WorkbookReader,
    WorkbookStructureError,
)

HEADERS = (
    "TT",
    "Họ và tên",
    "Ngày sinh",
    "Tổ",
    "Điểm gốc",
    "Điểm cộng",
    "Điểm trừ",
    "Tổng",
    "Hạnh kiểm",
    "Minh chứng",
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _create_supported_workbook(path: Path, *, stop_marker: str = "TỔNG HỢP") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Thi đua"
    worksheet["A1"] = "Đợt 6: Từ 28/02/2026 – 03/04/2026 (5 tuần)"
    worksheet["A2"] = "Bảng tổng hợp thử nghiệm"
    worksheet.append(HEADERS)
    worksheet.append(
        (
            1,
            "Học sinh A",
            date(2010, 1, 2),
            "1",
            10,
            5,
            1,
            14,
            "Tốt",
            "Đạt HĐTN 10/3(+5)",
        )
    )
    worksheet.append(
        (
            None,
            "Học sinh B",
            "03/04/2010",
            "2",
            10,
            "2,5",
            0,
            "=E5+F5-G5",
            "Tốt",
            "9đ Lí 13/3(+2,5)",
        )
    )
    worksheet.append(
        (
            3,
            "Học sinh C",
            40142,
            "3",
            10,
            0,
            0,
            9,
            "Khá",
            "4.8đ Toán GK2(-5)",
        )
    )
    worksheet.append((stop_marker,))
    worksheet.append((4, "Không được đọc", None, None, 10, 1, 0, 11, None, "(+1)"))
    workbook.save(path)
    _set_formula_cache(path, coordinate="H5", cached_value="12.5")


def _set_formula_cache(path: Path, *, coordinate: str, cached_value: str) -> None:
    replacement_path = path.with_suffix(".cached.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        replacement_path, "w"
    ) as destination:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                pattern = re.compile(
                    rb'(<c r="'
                    + coordinate.encode()
                    + rb'"[^>]*><f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)'
                )
                payload, count = pattern.subn(
                    rb"\g<1><v>" + cached_value.encode() + rb"</v>", payload, count=1
                )
                if count != 1:
                    raise AssertionError("formula cell was not found in worksheet XML")
            destination.writestr(item, payload)
    replacement_path.replace(path)


def test_reader_detects_header_period_students_formulas_and_mismatch(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    _create_supported_workbook(source)
    hash_before = _sha256(source)

    result = WorkbookReader().read(source)

    assert _sha256(source) == hash_before
    assert result.source_file_sha256 == hash_before
    assert result.sheet_name == "Thi đua"
    assert result.header_row == 3
    assert result.stopped_at_row == 7
    assert len(result.students) == 3
    assert {column.source_column for column in result.columns} == set(SourceColumn) - {
        SourceColumn.UNKNOWN
    }

    assert result.scoring_period is not None
    assert result.scoring_period.name == "Đợt 6"
    assert result.scoring_period.starts_on == date(2026, 2, 28)
    assert result.scoring_period.ends_on == date(2026, 4, 3)

    first, second, third = result.students
    assert first.excel_row == 4
    assert first.sequence_number == 1
    assert first.full_name_raw == "Học sinh A"
    assert first.birth_date == date(2010, 1, 2)
    assert first.base_score == Decimal("10")
    assert first.declared_positive_total == Decimal("5")
    assert first.declared_negative_total == Decimal("1")
    assert first.declared_final_total == Decimal("14")
    assert first.evidence_raw == "Đạt HĐTN 10/3(+5)"

    assert second.sequence_number is None
    assert second.full_name_raw == "Học sinh B"
    assert second.birth_date == date(2010, 4, 3)
    assert second.declared_positive_total == Decimal("2.5")
    assert second.declared_final_total == Decimal("12.5")
    total_cell = next(
        cell for cell in second.raw_row.cells if cell.source_column is SourceColumn.FINAL_TOTAL
    )
    assert total_cell.raw_text == "=E5+F5-G5"
    assert total_cell.formula == "=E5+F5-G5"
    assert total_cell.cached_value_text == "12.5"
    assert result.formulas_found is True
    assert result.cached_formula_values_found is True

    expected_serial_date = from_excel(40142).date()
    assert third.birth_date == expected_serial_date
    assert WarningCode.ROW_FORMULA_MISMATCH in {warning.code for warning in third.warnings}
    assert third.declared_final_total == Decimal("9")


@pytest.mark.parametrize("stop_marker", ["TỔNG HỢP", "Thành tích lớp", "GVCN"])
def test_reader_stops_at_non_student_sections(tmp_path: Path, stop_marker: str) -> None:
    source = tmp_path / f"stop-{stop_marker}.xlsx"
    _create_supported_workbook(source, stop_marker=stop_marker)

    result = WorkbookReader().read(source)

    assert result.stopped_at_row == 7
    assert [student.full_name_raw for student in result.students] == [
        "Học sinh A",
        "Học sinh B",
        "Học sinh C",
    ]


def test_reader_requires_exactly_one_sheet(tmp_path: Path) -> None:
    source = tmp_path / "two-sheets.xlsx"
    workbook = Workbook()
    workbook.create_sheet("Khác")
    workbook.save(source)

    with pytest.raises(WorkbookStructureError, match="đúng một sheet"):
        WorkbookReader().read(source)


def test_reader_rejects_missing_header(tmp_path: Path) -> None:
    source = tmp_path / "missing-header.xlsx"
    workbook = Workbook()
    workbook.active.append(("Họ và tên", "Minh chứng"))
    workbook.save(source)

    with pytest.raises(HeaderNotFoundError, match="Không tìm thấy"):
        WorkbookReader().read(source)


def test_reader_rejects_ambiguous_header(tmp_path: Path) -> None:
    source = tmp_path / "ambiguous-header.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(HEADERS)
    worksheet.append(HEADERS)
    workbook.save(source)

    with pytest.raises(HeaderAmbiguousError, match="Nhiều hàng"):
        WorkbookReader().read(source)


def test_reader_warns_but_keeps_numeric_tt_without_name(tmp_path: Path) -> None:
    source = tmp_path / "missing-name.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(HEADERS)
    worksheet.append((1, None, None, None, 10, 0, 0, 10, None, "Không có tên"))
    workbook.save(source)

    result = WorkbookReader().read(source)

    assert len(result.students) == 1
    assert result.students[0].full_name_raw is None
    assert WarningCode.MISSING_REQUIRED_COLUMN in {
        warning.code for warning in result.students[0].warnings
    }
